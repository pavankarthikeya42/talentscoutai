import os
import csv

# Test cases data list
TEST_CASES = [
    {
        "id": "TC-001",
        "module": "Auth & RBAC",
        "title": "Super Admin - Successful Login & Recruiter Management",
        "preconditions": "Super Admin account (superadmin@talentscout.ai) exists in Supabase DB.",
        "steps": "1. Open login page.\n2. Enter Super Admin credentials.\n3. Click Login.\n4. Navigate to Recruiters tab.",
        "expected": "User is authenticated and redirected to Dashboard. Recruiters list is visible with option to create/delete recruiters. Jobs/Candidates tabs are NOT visible or accessible.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-002",
        "module": "Auth & RBAC",
        "title": "Super Admin - Access Control for Jobs REST API",
        "preconditions": "Super Admin token generated.",
        "steps": "1. Send GET request to /api/jobs/ using Super Admin authorization token.",
        "expected": "HTTP 403 Forbidden or restricted response (SuperAdmin is restricted to recruiter management).",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-003",
        "module": "Auth & RBAC",
        "title": "HR Admin - Successful Login & Dashboard Access",
        "preconditions": "HR account (hr@talentscout.ai) exists in Supabase DB.",
        "steps": "1. Open login page.\n2. Enter HR credentials.\n3. Click Login.\n4. Verify sidebar tabs.",
        "expected": "User is authenticated. Sidebar shows Dashboard, Jobs, Candidates, Screening, Interviews, and AI Chat (RAG) tabs.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-004",
        "module": "Auth & RBAC",
        "title": "Manager - Successful Login & Access Scope",
        "preconditions": "Manager account (manager1@talentscout.ai) exists in Supabase DB.",
        "steps": "1. Open login page.\n2. Enter Manager credentials.\n3. Click Login.\n4. Verify sidebar tabs.",
        "expected": "User is authenticated. Sidebar shows Jobs, Candidates, and Interviews tabs. Recruiters and AI Chat (RAG) tabs are NOT visible.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-005",
        "module": "Auth & RBAC",
        "title": "Manager - API Isolation and Privilege Verification",
        "preconditions": "Manager token generated.",
        "steps": "1. Attempt to send GET to /api/auth/recruiters.\n2. Attempt to send POST to /api/rag/chat.",
        "expected": "Both requests return HTTP 403 Forbidden. Access is strictly denied to Manager role.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-006",
        "module": "Auth & RBAC",
        "title": "Public Candidate - Unauthorized Router Guard",
        "preconditions": "User is anonymous.",
        "steps": "1. Navigate directly to /dashboard or /candidates in the browser.",
        "expected": "User is redirected back to /login with state preserved.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-007",
        "module": "Auth & RBAC",
        "title": "Public Careers Portal Access",
        "preconditions": "Candidate careers portal has active job postings.",
        "steps": "1. Navigate to public careers root URL (e.g. /careers or /portal).\n2. Verify visibility of active jobs.",
        "expected": "Job listings are displayed correctly without prompting for authentication.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-008",
        "module": "Auth & RBAC",
        "title": "Candidate Application Submission",
        "preconditions": "Candidate on the public job application page.",
        "steps": "1. Select a job.\n2. Fill in Name, Email, and upload a resume (PDF).\n3. Click Submit Application.",
        "expected": "Application is successfully submitted. Candidate receives a confirmation UI message. Record is created in database.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-009",
        "module": "Auth & RBAC",
        "title": "Authentication - Invalid Credentials",
        "preconditions": "None.",
        "steps": "1. Navigate to login page.\n2. Enter invalid email or password.\n3. Click Login.",
        "expected": "Authentication fails. UI displays clean, user-friendly error: 'Invalid email or password'. No token is issued.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-010",
        "module": "Auth & RBAC",
        "title": "Authentication - JWT Token Expiration",
        "preconditions": "User logged in with a JWT token nearing expiration.",
        "steps": "1. Wait for token to expire.\n2. Attempt an authenticated request (e.g. refresh candidate list).",
        "expected": "System automatically attempts to refresh token or redirects to login page with clear feedback.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-011",
        "module": "Job Management",
        "title": "Manual Job Creation",
        "preconditions": "Logged in as Manager or HR.",
        "steps": "1. Navigate to Jobs tab -> Create Job.\n2. Enter Title, Department, Description, and Requirements.\n3. Click Create.",
        "expected": "Job is created successfully. Record in Supabase DB includes generated vector embedding for semantic matching. Job appears in the list.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-012",
        "module": "Job Management",
        "title": "AI Autofill Job Details",
        "preconditions": "Logged in as HR/Manager. Gemini/Groq API keys set in backend .env.",
        "steps": "1. Click 'Autofill Job Details' in the Job Form.\n2. Provide short prompt (e.g. 'Senior React Developer, 5 years exp, remote').\n3. Click Generate.",
        "expected": "AI generates rich Job Description, Roles & Responsibilities, and Technical Requirements. Fields in form autofill dynamically.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-013",
        "module": "Job Management",
        "title": "Edit Job Details",
        "preconditions": "Job exists and was created by current Manager (or user is HR).",
        "steps": "1. Select job -> Click Edit.\n2. Modify Requirements (e.g., append 'Docker').\n3. Save changes.",
        "expected": "Job details update in database. AI embedding is updated to reflect the new requirements.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-014",
        "module": "Job Management",
        "title": "Delete Job Cascading",
        "preconditions": "Job exists with associated candidate applications.",
        "steps": "1. Select job -> Click Delete -> Confirm.",
        "expected": "Job is soft-deleted or deleted from database. Connected applications are updated or deleted according to cascade rules.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-015",
        "module": "Job Management",
        "title": "Manager Job Isolation (Multi-Tenancy)",
        "preconditions": "Manager 1 and Manager 2 accounts exist; Manager 1 has Job A; Manager 2 has Job B.",
        "steps": "1. Log in as Manager 1. View Jobs list.\n2. Log in as Manager 2. View Jobs list.",
        "expected": "Manager 1 only sees Job A. Manager 2 only sees Job B. HR admin sees both Job A and Job B.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-016",
        "module": "Job Management",
        "title": "LinkedIn Integration Toggle",
        "preconditions": "None.",
        "steps": "1. Open job detail/edit form.\n2. Toggle 'Post to LinkedIn' on.\n3. Save Job.",
        "expected": "posted_to_linkedin flag is set to TRUE in the DB. Integration webhook or service mock is called.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-017",
        "module": "Job Management",
        "title": "Naukri Integration Toggle",
        "preconditions": "None.",
        "steps": "1. Open job detail/edit form.\n2. Toggle 'Post to Naukri' on.\n3. Save Job.",
        "expected": "posted_to_naukri flag is set to TRUE in the DB. Integration webhook or service mock is called.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-018",
        "module": "Resume Processing",
        "title": "Single Resume Upload (PDF)",
        "preconditions": "Logged in as HR Admin.",
        "steps": "1. Go to Resumes -> Upload.\n2. Select a valid PDF resume.\n3. Click Upload.",
        "expected": "Text is extracted from PDF. AI parses structural details (skills, experience, education). Candidate record and vector embedding are generated.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-019",
        "module": "Resume Processing",
        "title": "Single Resume Upload (DOCX)",
        "preconditions": "Logged in as HR Admin.",
        "steps": "1. Go to Resumes -> Upload.\n2. Select a valid DOCX resume.\n3. Click Upload.",
        "expected": "Text is extracted from DOCX. AI parses structural details successfully. Candidate record is saved with vector embedding.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-020",
        "module": "Resume Processing",
        "title": "Bulk Resume Upload (ZIP)",
        "preconditions": "ZIP file containing 5 distinct candidate resumes (PDFs and DOCXs).",
        "steps": "1. Go to Resumes -> Upload Bulk.\n2. Choose the ZIP file.\n3. Click Upload.",
        "expected": "ZIP file is unpacked asynchronously. All 5 resumes are processed in background. Notification or progress updates indicate success.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-021",
        "module": "Resume Processing",
        "title": "AI Parsing Quality and Mapping",
        "preconditions": "Resume with clear structure uploaded.",
        "steps": "1. Open candidate details page for the uploaded resume.\n2. Verify fields: Skills, Years of Experience, Education history.",
        "expected": "Skills are correctly classified. Experience timeline is parsed. Education fields are correctly normalized.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-022",
        "module": "Resume Processing",
        "title": "Resume Processing - Corrupted File Error Handling",
        "preconditions": "Corrupted or non-readable PDF file prepared.",
        "steps": "1. Attempt to upload the corrupted PDF.",
        "expected": "Backend catches extraction exception gracefully. Returns appropriate HTTP error (e.g. 422 Unprocessable Entity). UI displays readable error.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-023",
        "module": "Resume Processing",
        "title": "Trigger Global Embeddings Regeneration",
        "preconditions": "Super Admin or HR logged in.",
        "steps": "1. Click on Settings / Database Management.\n2. Click 'Re-embed All Candidates' button.",
        "expected": "Background task starts. Candidates table embeddings are updated. Success notification pops up once complete.",
        "priority": "Low",
        "status": "Converted"
    },
    {
        "id": "TC-024",
        "module": "AI Screening",
        "title": "Cosine Similarity Candidate Match",
        "preconditions": "Job opening with vector embedding. Matching and non-matching candidate profiles exist.",
        "steps": "1. Go to Screening tab -> Select Job.\n2. Click 'Screen Candidates'.",
        "expected": "Database performs cosine similarity search. Relevant candidates are returned, ordered by matching score (highest to lowest).",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-025",
        "module": "AI Screening",
        "title": "Candidate Suitability Scoring Logic",
        "preconditions": "Candidate matched against Job.",
        "steps": "1. Inspect suitability score details for candidate.",
        "expected": "Suitability score correctly reflects a hybrid calculation: cosine similarity + keyword hard-skill match + years of experience gap.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-026",
        "module": "AI Screening",
        "title": "Automated Email - Interview Invitation",
        "preconditions": "Candidate has status 'Shortlisted'. SMTP settings configured in backend .env.",
        "steps": "1. Select Shortlisted candidate.\n2. Click 'Send Automated Email'.\n3. Select template 'Interview Invitation'.\n4. Send.",
        "expected": "AI drafts a personalized email referencing candidate's specific skills. Email is sent successfully. Log recorded in database.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-027",
        "module": "AI Screening",
        "title": "Automated Email - Rejection",
        "preconditions": "Candidate has status 'Rejected'.",
        "steps": "1. Select Rejected candidate.\n2. Click 'Send Automated Email'.\n3. Select template 'Rejection'.\n4. Send.",
        "expected": "AI drafts a polite, personalized rejection email. Email is sent successfully. Log recorded.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-028",
        "module": "AI Screening",
        "title": "Application Status Transition",
        "preconditions": "Candidate application exists.",
        "steps": "1. Go to Screening -> Click status dropdown for application.\n2. Change from 'Applied' to 'Interviewing'.",
        "expected": "Status is updated in Supabase. Timeline or activity log is updated. UI displays the updated status instantly.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-029",
        "module": "Interview Management",
        "title": "Recommend Interview Rounds",
        "preconditions": "Candidate application active. Job description requirements exist.",
        "steps": "1. Open Candidate application details -> Interviews tab.\n2. Click 'Recommend Rounds'.",
        "expected": "AI analyzes candidate's skill gaps against Job requirements and returns specific proposed interview rounds (e.g. 'React Coding Challenge', 'System Design').",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-030",
        "module": "Interview Management",
        "title": "Generate AI Interview Questions",
        "preconditions": "Interview round scheduled.",
        "steps": "1. Click 'Generate Questions' for the scheduled round.",
        "expected": "AI generates 5 relevant, customized interview questions tailored to testing the candidate's specific experience and skills gap.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-031",
        "module": "Interview Management",
        "title": "Submit Interview Feedback",
        "preconditions": "Interviewer is logged in (Manager role).",
        "steps": "1. Select scheduled interview.\n2. Enter ratings (1-5), comments, and recommendation.\n3. Click Submit.",
        "expected": "Feedback is persisted. Average candidate rating is recalculated. HR is notified of feedback submission.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-032",
        "module": "RAG Chatbot",
        "title": "RAG Candidate Query Search",
        "preconditions": "Logged in as HR Admin. Candidates exist in DB.",
        "steps": "1. Open AI Chat.\n2. Type 'Find candidates who have python experience and have worked at Google'.\n3. Click Send.",
        "expected": "RAG agent performs vector search, formats candidate details, and answers the HR query naturally with direct links/references.",
        "priority": "High",
        "status": "Converted"
    },
    {
        "id": "TC-033",
        "module": "RAG Chatbot",
        "title": "Chat History Context Preservation",
        "preconditions": "Active chat session.",
        "steps": "1. Type 'Who is the most qualified Python developer?' -> Send.\n2. Receive response.\n3. Type 'What is their email address?' -> Send.",
        "expected": "AI understands 'their' refers to the Python developer mentioned in the previous turn and provides the correct email address.",
        "priority": "Medium",
        "status": "Converted"
    },
    {
        "id": "TC-034",
        "module": "End-to-End Flow",
        "title": "E2E Recruitment Funnel Walkthrough",
        "preconditions": "Full database set up, external APIs (Gemini/Groq/SMTP) mockable/available.",
        "steps": "1. Manager creates Job A.\n2. Candidate applies for Job A via careers page.\n3. HR screens candidates, AI scores Candidate A at 85%.\n4. HR changes status to Interviewing and schedules Python interview.\n5. Manager conducts interview and submits positive feedback.\n6. HR changes status to Hired.",
        "expected": "All stages transition smoothly. Database updates are immediate. User roles operate within designated scopes without crossing RBAC boundaries.",
        "priority": "High",
        "status": "Converted"
    }
]

def generate_csv(filepath):
    print(f"Generating CSV at: {filepath}...")
    headers = [
        "Test Case ID", 
        "Feature / Module", 
        "Test Case Description / Title", 
        "Preconditions", 
        "Steps to Execute", 
        "Expected Result", 
        "Priority", 
        "Status"
    ]
    with open(filepath, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for tc in TEST_CASES:
            writer.writerow([
                tc["id"],
                tc["module"],
                tc["title"],
                tc["preconditions"],
                tc["steps"],
                tc["expected"],
                tc["priority"],
                tc["status"]
            ])
    print("CSV generated successfully!")

def generate_xlsx(filepath):
    print("Attempting to generate styled Excel sheet using openpyxl...")
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is not installed. Styled Excel (.xlsx) file generation skipped.")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TalentScout AI Test Cases"
    
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Test Case ID", 
        "Feature / Module", 
        "Test Case Description / Title", 
        "Preconditions", 
        "Steps to Execute", 
        "Expected Result", 
        "Priority", 
        "Status"
    ]
    ws.append(headers)
    
    # Add data
    for tc in TEST_CASES:
        ws.append([
            tc["id"],
            tc["module"],
            tc["title"],
            tc["preconditions"],
            tc["steps"],
            tc["expected"],
            tc["priority"],
            tc["status"]
        ])
        
    # Styles definition
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    regular_font = Font(name=font_family, size=10)
    bold_font = Font(name=font_family, size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Priority and Status Fills
    priority_fills = {
        "High": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"), # Light Red/Orange
        "Medium": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), # Light Yellow
        "Low": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green
    }
    
    status_fills = {
        "Pending": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"), # Gray
        "Pass": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), # Soft Green
        "Fail": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), # Soft Red
        "Converted": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Soft Blue
    }
    
    # Format Headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        
    # Format Data Rows
    for row_idx in range(2, len(TEST_CASES) + 2):
        ws.row_dimensions[row_idx].height = 55 # Comfortable height for wraps
        
        # ID Alignment
        cell_id = ws.cell(row=row_idx, column=1)
        cell_id.alignment = Alignment(horizontal="center", vertical="center")
        cell_id.font = bold_font
        
        # Module alignment
        cell_mod = ws.cell(row=row_idx, column=2)
        cell_mod.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Other text alignments
        for col_idx in [3, 4, 5, 6]:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
        # Priority Highlight
        cell_pri = ws.cell(row=row_idx, column=7)
        cell_pri.alignment = Alignment(horizontal="center", vertical="center")
        pri_val = cell_pri.value
        if pri_val in priority_fills:
            cell_pri.fill = priority_fills[pri_val]
            cell_pri.font = bold_font
            
        # Status Highlight
        cell_stat = ws.cell(row=row_idx, column=8)
        cell_stat.alignment = Alignment(horizontal="center", vertical="center")
        stat_val = cell_stat.value
        if stat_val in status_fills:
            cell_stat.fill = status_fills[stat_val]
            cell_stat.font = bold_font
            
        # Apply borders and regular fonts where not styled
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx not in [1, 7, 8]:
                cell.font = regular_font

    # Adjust Column Widths nicely
    widths = {
        1: 15, # ID
        2: 18, # Module
        3: 35, # Title
        4: 35, # Preconditions
        5: 45, # Steps
        6: 45, # Expected
        7: 12, # Priority
        8: 12  # Status
    }
    for col_idx, width in widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    wb.save(filepath)
    print("XLSX generated successfully!")
    return True

if __name__ == "__main__":
    # Define paths
    root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    csv_path = os.path.join(root_dir, "test_cases.csv")
    xlsx_path = os.path.join(root_dir, "test_cases.xlsx")
    
    generate_csv(csv_path)
    
    # Try generating Excel. If not installed, it will print warning.
    success = generate_xlsx(xlsx_path)
    if not success:
        print("\nNote: Only CSV was generated because openpyxl is not installed in the current environment.")
        print("To generate a fully styled Excel (.xlsx) file, install openpyxl:")
        print("    pip install openpyxl")
        print("Then run this script again.")
